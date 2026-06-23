from datetime import datetime, timezone
import re
from pathlib import Path
from uuid import uuid4

from bson import ObjectId
from fastapi import BackgroundTasks, HTTPException, UploadFile, status
from openpyxl import load_workbook

from app.core.config import settings
from app.core.module_guard import ensure_tenant_module_enabled, ensure_tenant_module_usage_available
from app.core.object_ids import parse_object_id, serialize_document
from app.core.permissions import get_owned_tenant_or_403
from app.core.slug import slugify
from app.db.mongodb import get_database
from app.services.custom_field_service import validate_custom_values
from app.services.rag_index_service import index_item_for_rag
from app.services.storage_service import store_item_image
from app.services.phase32_utils import normalize_excel_header

ALLOWED_ITEM_TYPES = {"product", "service", "package", "raw_material", "asset", "digital_product"}
ALLOWED_ITEM_STATUSES = {"active", "inactive", "archived"}
ALLOWED_UNITS = {"piece", "hour", "session", "kg", "liter", "month", "custom"}
ALLOWED_SERVICE_DELIVERY_MODES = {"onsite", "remote", "pickup", "home_visit", "hybrid"}


async def _ensure_item_access(tenant_id: str, user: dict) -> ObjectId:
    tenant_oid = parse_object_id(tenant_id, "tenantId")
    await get_owned_tenant_or_403(tenant_oid, user)
    await ensure_tenant_module_enabled(tenant_oid, "items")
    return tenant_oid


async def _ensure_category_exists(tenant_oid: ObjectId, category_id: str | None) -> ObjectId | None:
    if not category_id:
        return None
    db = get_database()
    category_oid = parse_object_id(category_id, "categoryId")
    category = await db.item_categories.find_one({"_id": category_oid, "tenantId": tenant_oid})
    if not category:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item category not found.")
    return category_oid


def _stock_dict(stock) -> dict:
    return {
        "quantity": stock.quantity,
        "lowStockThreshold": stock.lowStockThreshold,
        "reservedQuantity": stock.reservedQuantity,
    }


def _image_dicts(images) -> list[dict]:
    return [image.model_dump() for image in images]


def _service_details_dict(service_details) -> dict:
    return {
        "durationMinutes": service_details.durationMinutes,
        "bufferMinutes": service_details.bufferMinutes,
        "deliveryMode": service_details.deliveryMode,
    }


def _variant_dicts(variants) -> list[dict]:
    return [
        {
            "name": variant.name.strip(),
            "sku": variant.sku.strip(),
            "price": variant.price,
            "compareAtPrice": variant.compareAtPrice,
            "stockQuantity": variant.stockQuantity,
            "lowStockThreshold": variant.lowStockThreshold,
            "isDefault": variant.isDefault,
            "isActive": variant.isActive,
            "optionValues": dict(variant.optionValues),
        }
        for variant in variants
    ]


def _bundle_component_dicts(bundle_components, component_items: list[dict]) -> list[dict]:
    rows = []
    for component, component_item in zip(bundle_components, component_items, strict=False):
        rows.append(
            {
                "itemId": component_item["_id"],
                "itemName": component_item.get("name", ""),
                "itemType": component_item.get("itemType", ""),
                "quantity": component.quantity,
                "isOptional": component.isOptional,
                "notes": component.notes,
            }
        )
    return rows


def _normalize_tags(tags: list[str] | None) -> list[str]:
    normalized = []
    seen = set()
    for tag in tags or []:
        clean = " ".join(str(tag).strip().split())
        if not clean:
            continue
        lowered = clean.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(clean)
    return normalized


def _validate_service_details(item_type: str, is_bookable: bool, service_details: dict) -> None:
    delivery_mode = str(service_details.get("deliveryMode") or "onsite").strip().lower()
    if delivery_mode not in ALLOWED_SERVICE_DELIVERY_MODES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid service delivery mode.")

    duration = int(service_details.get("durationMinutes", 0) or 0)
    buffer_minutes = int(service_details.get("bufferMinutes", 0) or 0)
    if duration < 0 or buffer_minutes < 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Service duration values cannot be negative.")
    if (item_type == "service" or is_bookable) and duration <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Services and bookable items require a duration in minutes.")


def _validate_variants(variants: list[dict], item_price: float) -> None:
    if not variants:
        return

    seen_names = set()
    seen_skus = set()
    default_count = 0
    for variant in variants:
        name = variant["name"].strip()
        if not name:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Variant name is required.")
        lowered_name = name.lower()
        if lowered_name in seen_names:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Duplicate variant name '{name}'.")
        seen_names.add(lowered_name)

        sku = variant["sku"].strip()
        if sku:
            lowered_sku = sku.lower()
            if lowered_sku in seen_skus:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Duplicate variant SKU '{sku}'.")
            seen_skus.add(lowered_sku)

        if variant["compareAtPrice"] is not None and variant["compareAtPrice"] < variant["price"]:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Compare-at price for '{name}' cannot be below the variant price.")
        if variant["stockQuantity"] < 0 or variant["lowStockThreshold"] < 0:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Stock values for '{name}' cannot be negative.")

        if variant["isDefault"]:
            default_count += 1

    if default_count > 1:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Only one variant can be marked as default.")
    if default_count == 0:
        variants[0]["isDefault"] = True

    if all(variant["price"] == 0 for variant in variants) and item_price <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Add a price to the item or at least one variant.")


async def _resolve_bundle_component_ids(
    tenant_oid: ObjectId,
    item_type: str,
    bundle_components,
    current_item_oid: ObjectId | None = None,
) -> list[dict]:
    if not bundle_components:
        return []
    if item_type != "package":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Bundle components are only supported for package items.")

    db = get_database()
    component_ids = []
    for component in bundle_components:
        component_oid = parse_object_id(component.itemId, "bundle itemId")
        if current_item_oid and component_oid == current_item_oid:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="An item cannot include itself in its own bundle.")
        component_ids.append(component_oid)

    unique_ids = list(dict.fromkeys(component_ids))
    items = await db.items.find({"_id": {"$in": unique_ids}, "tenantId": tenant_oid, "status": {"$ne": "archived"}}).to_list(length=len(unique_ids))
    if len(items) != len(unique_ids):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="One or more bundle items were not found.")
    item_map = {item["_id"]: item for item in items}
    return [item_map[component_oid] for component_oid in component_ids]


async def _validate_item_payload(tenant_id: str, tenant_oid: ObjectId, payload, user: dict, current_item_oid: ObjectId | None = None) -> dict:
    item_type = getattr(payload, "itemType", None)
    item_status = getattr(payload, "status", None)
    unit = getattr(payload, "unit", None)

    if item_type and item_type not in ALLOWED_ITEM_TYPES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item type.")
    if item_status and item_status not in ALLOWED_ITEM_STATUSES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item status.")
    if unit and unit not in ALLOWED_UNITS:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item unit.")

    category_oid = await _ensure_category_exists(tenant_oid, getattr(payload, "categoryId", None))
    custom_fields = getattr(payload, "customFields", None) or {}
    validation = await validate_custom_values(tenant_id, "items", "item", custom_fields, user)
    if not validation["valid"]:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=validation["errors"])

    service_details = _service_details_dict(getattr(payload, "serviceDetails"))
    is_bookable = bool(getattr(payload, "isBookable", False))
    _validate_service_details(str(item_type or "product"), is_bookable, service_details)

    variants = _variant_dicts(getattr(payload, "variants", []))
    _validate_variants(variants, float(getattr(payload, "price", 0) or 0))

    bundle_component_items = await _resolve_bundle_component_ids(tenant_oid, str(item_type or "product"), getattr(payload, "bundleComponents", []), current_item_oid)
    bundle_components = _bundle_component_dicts(getattr(payload, "bundleComponents", []), bundle_component_items)

    return {
        "categoryId": category_oid,
        "customFields": validation["values"],
        "serviceDetails": service_details,
        "variants": variants,
        "bundleComponents": bundle_components,
    }


async def create_item_category(tenant_id: str, payload, user: dict) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    parent_oid = await _ensure_category_exists(tenant_oid, payload.parentCategoryId)
    base_slug = slugify(payload.name)
    slug = base_slug
    counter = 2
    while await db.item_categories.find_one({"tenantId": tenant_oid, "slug": slug}):
        slug = f"{base_slug}-{counter}"
        counter += 1

    now = datetime.now(timezone.utc)
    category = {
        "tenantId": tenant_oid,
        "name": payload.name,
        "slug": slug,
        "description": payload.description,
        "parentCategoryId": parent_oid,
        "isActive": payload.isActive,
        "createdAt": now,
        "updatedAt": now,
    }
    category["_id"] = (await db.item_categories.insert_one(category)).inserted_id
    return serialize_document(category)


async def list_item_categories(tenant_id: str, user: dict, active_only: bool = False) -> list[dict]:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    query = {"tenantId": tenant_oid}
    if active_only:
        query["isActive"] = True
    cursor = db.item_categories.find(query).sort([("name", 1)])
    return [serialize_document(category) async for category in cursor]


async def update_item_category(tenant_id: str, category_id: str, payload, user: dict) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    category_oid = parse_object_id(category_id, "categoryId")
    existing = await db.item_categories.find_one({"_id": category_oid, "tenantId": tenant_oid})
    if not existing:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item category not found.")

    update = {"updatedAt": datetime.now(timezone.utc)}
    for key in ["name", "description", "isActive"]:
        value = getattr(payload, key)
        if value is not None:
            update[key] = value
    if payload.parentCategoryId is not None:
        update["parentCategoryId"] = await _ensure_category_exists(tenant_oid, payload.parentCategoryId)

    await db.item_categories.update_one({"_id": category_oid, "tenantId": tenant_oid}, {"$set": update})
    return serialize_document(await db.item_categories.find_one({"_id": category_oid, "tenantId": tenant_oid}))


async def delete_item_category(tenant_id: str, category_id: str, user: dict) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    category_oid = parse_object_id(category_id, "categoryId")
    result = await db.item_categories.update_one(
        {"_id": category_oid, "tenantId": tenant_oid},
        {"$set": {"isActive": False, "updatedAt": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item category not found.")
    return serialize_document(await db.item_categories.find_one({"_id": category_oid, "tenantId": tenant_oid}))


async def create_item(tenant_id: str, payload, user: dict, background_tasks: BackgroundTasks | None = None) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    await ensure_tenant_module_usage_available(tenant_oid, "items")
    validated = await _validate_item_payload(tenant_id, tenant_oid, payload, user)
    now = datetime.now(timezone.utc)
    item = {
        "tenantId": tenant_oid,
        "branchId": None,
        "itemType": payload.itemType,
        "name": payload.name,
        "description": payload.description,
        "categoryId": validated["categoryId"],
        "sku": payload.sku,
        "price": payload.price,
        "costPrice": payload.costPrice,
        "currency": payload.currency,
        "unit": payload.unit,
        "images": _image_dicts(payload.images),
        "status": payload.status,
        "isSellable": payload.isSellable,
        "isBookable": payload.isBookable,
        "isStockTracked": payload.isStockTracked,
        "stock": _stock_dict(payload.stock),
        "serviceDetails": validated["serviceDetails"],
        "variants": validated["variants"],
        "bundleComponents": validated["bundleComponents"],
        "customFields": validated["customFields"],
        "tags": _normalize_tags(payload.tags),
        "createdAt": now,
        "updatedAt": now,
    }
    item["_id"] = (await db.items.insert_one(item)).inserted_id
    if background_tasks:
        background_tasks.add_task(index_item_for_rag, tenant_oid, item)
    else:
        await index_item_for_rag(tenant_oid, item)
    return serialize_document(item)


async def list_items(
    tenant_id: str,
    user: dict,
    search: str = "",
    status_filter: str | None = None,
    item_type: str | None = None,
    category_id: str | None = None,
    page: int = 1,
    limit: int = 20,
) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    page = max(page, 1)
    limit = min(max(limit, 1), 100)
    query = {"tenantId": tenant_oid}
    if status_filter:
        query["status"] = status_filter
    if item_type:
        query["itemType"] = item_type
    if category_id:
        query["categoryId"] = parse_object_id(category_id, "categoryId")
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"tags": {"$regex": search, "$options": "i"}},
            {"variants.name": {"$regex": search, "$options": "i"}},
            {"variants.sku": {"$regex": search, "$options": "i"}},
        ]

    total = await db.items.count_documents(query)
    cursor = db.items.find(query).sort("createdAt", -1).skip((page - 1) * limit).limit(limit)
    return {
        "items": [serialize_document(item) async for item in cursor],
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "totalPages": (total + limit - 1) // limit,
        },
    }


async def get_item(tenant_id: str, item_id: str, user: dict) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    item_oid = parse_object_id(item_id, "itemId")
    item = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found.")
    return serialize_document(item)


async def update_item(tenant_id: str, item_id: str, payload, user: dict, background_tasks: BackgroundTasks | None = None) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    item_oid = parse_object_id(item_id, "itemId")
    existing = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if not existing:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found.")

    candidate_item_type = payload.itemType or existing.get("itemType", "product")
    candidate_status = payload.status or existing.get("status", "active")
    candidate_unit = payload.unit or existing.get("unit", "piece")
    if candidate_item_type not in ALLOWED_ITEM_TYPES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item type.")
    if candidate_status not in ALLOWED_ITEM_STATUSES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item status.")
    if candidate_unit not in ALLOWED_UNITS:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid item unit.")

    merged_payload = type(
        "MergedItemPayload",
        (),
        {
            "itemType": candidate_item_type,
            "status": candidate_status,
            "unit": candidate_unit,
            "categoryId": payload.categoryId if payload.categoryId is not None else (str(existing.get("categoryId")) if existing.get("categoryId") else None),
            "customFields": payload.customFields if payload.customFields is not None else existing.get("customFields", {}),
            "serviceDetails": payload.serviceDetails if payload.serviceDetails is not None else type("ServiceDetailsPayload", (), existing.get("serviceDetails", {"durationMinutes": 0, "bufferMinutes": 0, "deliveryMode": "onsite"}))(),
            "variants": payload.variants if payload.variants is not None else [type("VariantPayload", (), variant)() for variant in existing.get("variants", [])],
            "bundleComponents": payload.bundleComponents if payload.bundleComponents is not None else [type("BundlePayload", (), {**component, "itemId": str(component["itemId"])})() for component in existing.get("bundleComponents", [])],
            "isBookable": payload.isBookable if payload.isBookable is not None else existing.get("isBookable", False),
            "price": payload.price if payload.price is not None else existing.get("price", 0),
        },
    )()
    validated = await _validate_item_payload(tenant_id, tenant_oid, merged_payload, user, item_oid)

    update = {"updatedAt": datetime.now(timezone.utc)}
    for key in ["itemType", "name", "description", "sku", "price", "costPrice", "currency", "unit", "status", "isSellable", "isBookable", "isStockTracked"]:
        value = getattr(payload, key)
        if value is not None:
            update[key] = value
    if payload.tags is not None:
        update["tags"] = _normalize_tags(payload.tags)
    if payload.categoryId is not None:
        update["categoryId"] = validated["categoryId"]
    if payload.images is not None:
        update["images"] = _image_dicts(payload.images)
    if payload.stock is not None:
        update["stock"] = _stock_dict(payload.stock)
    if payload.customFields is not None:
        update["customFields"] = validated["customFields"]
    if payload.serviceDetails is not None:
        update["serviceDetails"] = validated["serviceDetails"]
    if payload.variants is not None:
        update["variants"] = validated["variants"]
    if payload.bundleComponents is not None:
        update["bundleComponents"] = validated["bundleComponents"]

    await db.items.update_one({"_id": item_oid, "tenantId": tenant_oid}, {"$set": update})
    item = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if background_tasks:
        background_tasks.add_task(index_item_for_rag, tenant_oid, item)
    else:
        await index_item_for_rag(tenant_oid, item)
    return serialize_document(item)


async def delete_item(tenant_id: str, item_id: str, user: dict, background_tasks: BackgroundTasks | None = None) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    item_oid = parse_object_id(item_id, "itemId")
    result = await db.items.update_one(
        {"_id": item_oid, "tenantId": tenant_oid},
        {"$set": {"status": "archived", "updatedAt": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found.")
    item = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if background_tasks:
        background_tasks.add_task(index_item_for_rag, tenant_oid, item)
    else:
        await index_item_for_rag(tenant_oid, item)
    return serialize_document(item)


async def upload_item_image(tenant_id: str, item_id: str, file: UploadFile, user: dict, background_tasks: BackgroundTasks | None = None) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    item_oid = parse_object_id(item_id, "itemId")
    item = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found.")

    image = await store_item_image(tenant_id, item_id, file)
    await db.items.update_one(
        {"_id": item_oid, "tenantId": tenant_oid},
        {"$push": {"images": image}, "$set": {"updatedAt": datetime.now(timezone.utc)}},
    )
    item = await db.items.find_one({"_id": item_oid, "tenantId": tenant_oid})
    if background_tasks:
        background_tasks.add_task(index_item_for_rag, tenant_oid, item)
    return serialize_document(item)


def _bool_value(value, default=False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _float_value(value, default=0) -> float:
    if value is None or value == "":
        return default
    return float(value)


def _int_value(value, default=0) -> int:
    if value is None or value == "":
        return default
    return int(float(value))


async def import_items_from_excel(tenant_id: str, file: UploadFile, user: dict, background_tasks: BackgroundTasks | None = None) -> dict:
    db = get_database()
    tenant_oid = await _ensure_item_access(tenant_id, user)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Upload an .xlsx or .xlsm Excel file.")

    temp_dir = Path(settings.temp_upload_dir).resolve()
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / f"{uuid4().hex}_{file.filename}"
    try:
        with temp_path.open("wb") as output:
            while chunk := await file.read(1024 * 1024):
                output.write(chunk)

        workbook = load_workbook(temp_path)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Excel file is empty.")

        raw_headers = [str(header).strip() if header is not None else "" for header in rows[0]]
        headers = [normalize_excel_header(header) for header in raw_headers]
        header_map = {header: index for index, header in enumerate(headers) if header}
        errors = []
        created_items = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(cell not in (None, "") for cell in row):
                continue
            try:
                row_data = {header: row[index] if index < len(row) else None for header, index in header_map.items()}
                category_oid = None
                category_value = str(row_data.get("category") or "").strip()
                if category_value:
                    category = await db.item_categories.find_one(
                        {"tenantId": tenant_oid, "name": {"$regex": f"^{re.escape(category_value)}$", "$options": "i"}}
                    )
                    if not category:
                        category_slug_base = slugify(category_value)
                        category_slug = category_slug_base
                        category_counter = 2
                        while await db.item_categories.find_one({"tenantId": tenant_oid, "slug": category_slug}):
                            category_slug = f"{category_slug_base}-{category_counter}"
                            category_counter += 1
                        category = {
                            "tenantId": tenant_oid,
                            "name": category_value,
                            "slug": category_slug,
                            "description": "Created automatically from Excel import.",
                            "parentCategoryId": None,
                            "isActive": True,
                            "createdAt": datetime.now(timezone.utc),
                            "updatedAt": datetime.now(timezone.utc),
                        }
                        category["_id"] = (await db.item_categories.insert_one(category)).inserted_id
                    category_oid = category["_id"]

                item_type = str(row_data.get("itemType") or "product").strip()
                status_value = str(row_data.get("status") or "active").strip()
                unit_value = str(row_data.get("unit") or "piece").strip()
                if item_type not in ALLOWED_ITEM_TYPES:
                    raise ValueError("Invalid itemType.")
                if status_value not in ALLOWED_ITEM_STATUSES:
                    raise ValueError("Invalid status.")
                if unit_value not in ALLOWED_UNITS:
                    raise ValueError("Invalid unit.")

                custom_values = {
                    header.replace("custom.", "", 1): row_data.get(header)
                    for header in headers
                    if header.startswith("custom.")
                }
                validation = await validate_custom_values(tenant_id, "items", "item", custom_values, user)
                if not validation["valid"]:
                    raise ValueError("; ".join(error["message"] for error in validation["errors"]))

                service_details = {
                    "durationMinutes": _int_value(row_data.get("serviceDurationMinutes")),
                    "bufferMinutes": _int_value(row_data.get("serviceBufferMinutes")),
                    "deliveryMode": str(row_data.get("serviceDeliveryMode") or "onsite").strip().lower(),
                }
                _validate_service_details(item_type, _bool_value(row_data.get("isBookable"), False), service_details)

                now = datetime.now(timezone.utc)
                item = {
                    "tenantId": tenant_oid,
                    "branchId": None,
                    "itemType": item_type,
                    "name": str(row_data.get("name") or "").strip(),
                    "description": str(row_data.get("description") or ""),
                    "categoryId": category_oid,
                    "sku": str(row_data.get("sku") or ""),
                    "price": _float_value(row_data.get("price")),
                    "costPrice": _float_value(row_data.get("costPrice")),
                    "currency": str(row_data.get("currency") or "PKR"),
                    "unit": unit_value,
                    "images": ([{"provider": "external", "fileId": str(row_data.get("imageUrl") or "").strip(), "url": str(row_data.get("imageUrl") or "").strip()}] if str(row_data.get("imageUrl") or "").strip() else []),
                    "status": status_value,
                    "isSellable": _bool_value(row_data.get("isSellable"), True),
                    "isBookable": _bool_value(row_data.get("isBookable"), False),
                    "isStockTracked": _bool_value(row_data.get("isStockTracked"), True),
                    "stock": {
                        "quantity": _float_value(row_data.get("stockQuantity")),
                        "lowStockThreshold": _float_value(row_data.get("lowStockThreshold")),
                        "reservedQuantity": 0,
                    },
                    "serviceDetails": service_details,
                    "variants": [],
                    "bundleComponents": [],
                    "customFields": validation["values"],
                    "tags": [tag.strip() for tag in str(row_data.get("tags") or "").split(",") if tag.strip()],
                    "createdAt": now,
                    "updatedAt": now,
                }
                option_values = {}
                for option_key in ["color", "size", "material"]:
                    option_value = str(row_data.get(option_key) or "").strip()
                    if option_value:
                        option_values[option_key.title()] = option_value
                if option_values:
                    variant_name = " / ".join(option_values.values())
                    item["variants"] = [
                        {
                            "name": variant_name,
                            "sku": item["sku"],
                            "price": item["price"],
                            "compareAtPrice": None,
                            "stockQuantity": item["stock"]["quantity"],
                            "reservedQuantity": 0,
                            "lowStockThreshold": item["stock"]["lowStockThreshold"],
                            "isDefault": True,
                            "isActive": True,
                            "optionValues": option_values,
                        }
                    ]
                if len(item["name"]) < 2:
                    raise ValueError("Name is required.")
                if item["price"] < 0 or item["costPrice"] < 0:
                    raise ValueError("Prices cannot be negative.")
                item["_id"] = (await db.items.insert_one(item)).inserted_id
                created_items.append(item)
                if background_tasks:
                    background_tasks.add_task(index_item_for_rag, tenant_oid, item)
                else:
                    await index_item_for_rag(tenant_oid, item)
            except Exception as exc:
                errors.append({"row": row_number, "message": str(exc)})

        import_record = {
            "tenantId": tenant_oid,
            "fileName": file.filename,
            "totalRows": max(len(rows) - 1, 0),
            "successCount": len(created_items),
            "errorCount": len(errors),
            "errors": errors,
            "status": "completed_with_errors" if errors else "completed",
            "createdBy": user["_id"],
            "createdAt": datetime.now(timezone.utc),
        }
        import_record["_id"] = (await db.item_imports.insert_one(import_record)).inserted_id
        return {
            "import": serialize_document(import_record),
            "items": [serialize_document(item) for item in created_items],
        }
    finally:
        temp_path.unlink(missing_ok=True)
